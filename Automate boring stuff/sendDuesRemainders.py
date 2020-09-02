# python 3
# program to send remainder email based on payment status of spreadsheet

import os
import openpyxl
import sys
import smtplib

os.system('cls')

#open spreadsheet and get latest due status

wb = openpyxl.load_workbook('C:/Users/NagaVenkataSuryaNare/Documents/Naresh/automate_online-materials/duesRecords.xlsx')
sheet =  wb['Sheet1']
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

# check each member's payment status
unpaidMembers = {}
for r in range(2, sheet.max_row+ 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email

# Login into email account
useremailid = input('Enter your login id: ')
passwrd = input('enter ur login pass: ')
smtpObj =  smtplib.SMTP('smtp.gmail.com', 587)
smtpObj.ehlo()
smtpObj.starttls()
smtpObj.login(useremailid, passwrd)

#send out remainder emails 
for name, email in unpaidMembers.items():
    body = "subject: %s dues unpaid. \nDear %s. \nRecords indicate that you have not paid dues for %s.    \
    Please make this payment as soo as possible. Thank you!'"     %(latestMonth, name, latestMonth)
    print('sending email to %s.....' % email)
    sendmailStatus = smtpObj.sendmail(useremailid, email, body)

    if sendmailStatus != {}:
        print('there was a problem sending email to %s: %s' %(email, sendmailStatus))
    smtpObj.quit()
